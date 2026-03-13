from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def style_header(ws, row=1):
    fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for cell in ws[row]:
        if cell.value is None:
            continue
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def autofit(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


wb = Workbook()
ws = wb.active
ws.title = "Core_Assumptions"

# Core assumptions
ws.append(["Variable", "Value", "Justification"])
rows = [
    ("Currency", "CAD", "Model in Canadian dollars"),
    ("Monthly ARPU Y1-Y5", "244,259,274,287,299", "Plan mix shifts toward Professional and Enterprise"),
    ("Monthly churn Y1-Y5", "1.8%,1.6%,1.4%,1.4%,1.4%", "Retention improves with workflow embedding"),
    ("Avg providers per clinic", 4.2, "Small-medium outpatient clinic target"),
    ("COGS per provider per month", 54, "Inference + hosting + storage + support + fees"),
    ("Implementation fee per new clinic Y1-Y5", "1500,1500,2000,2000,2000", "Onboarding and training"),
    ("Capex Y1-Y5", "180000,220000,280000,350000,420000", "Security and platform scaling"),
    ("Initial cash", 200000, "Founder and pre-seed cash"),
    ("Financing Y1-Y5", "1500000,1800000,2000000,0,0", "Fundraise plan to break-even"),
]
for r in rows:
    ws.append(r)
style_header(ws)
autofit(ws, {"A": 44, "B": 28, "C": 70})

# Growth and pricing
ws2 = wb.create_sheet("Growth_Pricing")
ws2.append(
    [
        "Year",
        "Clinics_EOY",
        "Providers_EOY",
        "Providers_Avg",
        "Basic_Mix",
        "Pro_Mix",
        "Ent_Mix",
        "Basic_Price",
        "Pro_Price",
        "Ent_Price",
        "Blended_ARPU",
        "Sub_Revenue",
        "Impl_Fee",
        "Impl_Revenue",
        "Total_Revenue",
    ]
)
data = [
    ("Y1", 22, 90, 70, 0.60, 0.35, 0.05, 199, 299, 399, None, None, 1500, None, None),
    ("Y2", 76, 320, 260, 0.50, 0.40, 0.10, 199, 299, 399, None, None, 1500, None, None),
    ("Y3", 214, 900, 760, 0.40, 0.45, 0.15, 199, 299, 399, None, None, 2000, None, None),
    ("Y4", 500, 2100, 1850, 0.32, 0.48, 0.20, 199, 299, 399, None, None, 2000, None, None),
    ("Y5", 980, 4200, 3700, 0.25, 0.50, 0.25, 199, 299, 399, None, None, 2000, None, None),
]
for r in data:
    ws2.append(list(r))

for row in range(2, 7):
    ws2[f"K{row}"] = f"=E{row}*H{row}+F{row}*I{row}+G{row}*J{row}"
    ws2[f"L{row}"] = f"=D{row}*K{row}*12"
    ws2[f"N{row}"] = f"=B{row}*M{row}"
    ws2[f"O{row}"] = f"=L{row}+N{row}"

style_header(ws2)
autofit(
    ws2,
    {
        "A": 8,
        "B": 12,
        "C": 14,
        "D": 14,
        "E": 10,
        "F": 10,
        "G": 10,
        "H": 12,
        "I": 10,
        "J": 10,
        "K": 12,
        "L": 14,
        "M": 10,
        "N": 13,
        "O": 13,
    },
)

# COGS
ws3 = wb.create_sheet("COGS")
ws3.append(["Component", "Cost_per_provider_per_month"])
for r in [
    ("LLM inference + summarization", 24),
    ("Cloud hosting + networking", 11),
    ("Storage + logging", 5),
    ("Support + implementation servicing", 8),
    ("Payment processing + marketplace fees", 6),
    ("Total COGS per provider/month", "=SUM(B2:B6)"),
]:
    ws3.append(r)
style_header(ws3)
ws3.append([])
ws3.append(["Year", "Avg_Providers", "Revenue", "COGS", "Gross_Profit", "Gross_Margin"])
for idx, year_row in enumerate(range(2, 7), start=10):
    ws3[f"A{idx}"] = f"Y{idx-9}"
    ws3[f"B{idx}"] = f"=Growth_Pricing!D{year_row}"
    ws3[f"C{idx}"] = f"=Growth_Pricing!O{year_row}"
    ws3[f"D{idx}"] = f"=B{idx}*$B$7*12"
    ws3[f"E{idx}"] = f"=C{idx}-D{idx}"
    ws3[f"F{idx}"] = f"=E{idx}/C{idx}"
style_header(ws3, row=9)
autofit(ws3, {"A": 38, "B": 24, "C": 15, "D": 14, "E": 14, "F": 14})

# Hiring plan
ws4 = wb.create_sheet("Hiring")
ws4.append(["Role", "Salary", "Y1", "Y2", "Y3", "Y4", "Y5", "Year5_Annual_Cost"])
hiring_rows = [
    ("Leadership/Ops", 180000, 2, 2, 2, 2, 2),
    ("Backend/Platform Engineer", 135000, 2, 2, 5, 6, 7),
    ("ML Engineer", 145000, 1, 2, 3, 3, 4),
    ("Clinical AI Specialist", 125000, 1, 1, 1, 2, 2),
    ("Product Manager", 140000, 0, 1, 1, 2, 3),
    ("Compliance/Privacy Officer", 120000, 0, 1, 1, 1, 2),
    ("Sales/Partnerships (OTE)", 130000, 0, 0, 1, 3, 6),
    ("Customer Success/Implementation", 95000, 0, 1, 2, 5, 7),
]
for idx, r in enumerate(hiring_rows, start=2):
    ws4.append(list(r) + [f"=B{idx}*G{idx}"])

style_header(ws4)
ws4.append([])
ws4.append(["Metric", "Y1", "Y2", "Y3", "Y4", "Y5"])
ws4.append(["Total FTE", "=SUM(C2:C9)", "=SUM(D2:D9)", "=SUM(E2:E9)", "=SUM(F2:F9)", "=SUM(G2:G9)"])
ws4.append(
    [
        "Base Payroll",
        "=SUMPRODUCT(B2:B9,C2:C9)",
        "=SUMPRODUCT(B2:B9,D2:D9)",
        "=SUMPRODUCT(B2:B9,E2:E9)",
        "=SUMPRODUCT(B2:B9,F2:F9)",
        "=SUMPRODUCT(B2:B9,G2:G9)",
    ]
)
ws4.append(["Payroll w/18% burden", "=B12*1.18", "=C12*1.18", "=D12*1.18", "=E12*1.18", "=F12*1.18"])
style_header(ws4, row=11)
autofit(ws4, {"A": 36, "B": 12, "C": 10, "D": 10, "E": 10, "F": 10, "G": 10, "H": 16})

# Opex
ws5 = wb.create_sheet("Opex")
ws5.append(["Category", "Y1", "Y2", "Y3", "Y4", "Y5"])
opex_rows = [
    ("Marketing and demand generation", 70000, 120000, 220000, 620000, 1200000),
    ("Legal and compliance", 35000, 45000, 70000, 120000, 200000),
    ("Conferences and travel", 12000, 18000, 35000, 70000, 120000),
    ("SaaS tools and software ops", 28000, 40000, 65000, 130000, 210000),
    ("Infrastructure overhead (non-COGS)", 25000, 45000, 85000, 180000, 250000),
    ("Insurance and admin overhead", 18000, 30000, 58000, 98000, 117000),
]
for r in opex_rows:
    ws5.append(r)
ws5.append(
    [
        "Total non-payroll Opex",
        "=SUM(B2:B7)",
        "=SUM(C2:C7)",
        "=SUM(D2:D7)",
        "=SUM(E2:E7)",
        "=SUM(F2:F7)",
    ]
)
style_header(ws5)
autofit(ws5, {"A": 40, "B": 12, "C": 12, "D": 12, "E": 12, "F": 12})

# Financial statements
ws6 = wb.create_sheet("Financials")
ws6.append(["Metric", "Y1", "Y2", "Y3", "Y4", "Y5"])
financial_rows = [
    ("Revenue", "=Growth_Pricing!O2", "=Growth_Pricing!O3", "=Growth_Pricing!O4", "=Growth_Pricing!O5", "=Growth_Pricing!O6"),
    ("COGS", "=COGS!D10", "=COGS!D11", "=COGS!D12", "=COGS!D13", "=COGS!D14"),
    ("Gross Profit", "=B2-B3", "=C2-C3", "=D2-D3", "=E2-E3", "=F2-F3"),
    (
        "Payroll SG&A",
        "=Hiring!B13",
        "=Hiring!C13",
        "=Hiring!D13",
        "=Hiring!E13",
        "=Hiring!F13",
    ),
    ("Non-Payroll Opex", "=Opex!B8", "=Opex!C8", "=Opex!D8", "=Opex!E8", "=Opex!F8"),
    ("Total SG&A", "=B5+B6", "=C5+C6", "=D5+D6", "=E5+E6", "=F5+F6"),
    ("EBITDA", "=B4-B7", "=C4-C7", "=D4-D7", "=E4-E7", "=F4-F7"),
    ("EBITDA Margin", "=B8/B2", "=C8/C2", "=D8/D2", "=E8/E2", "=F8/F2"),
    ("Capex", 180000, 220000, 280000, 350000, 420000),
    ("Net Cash Flow pre-financing", "=B8-B10", "=C8-C10", "=D8-D10", "=E8-E10", "=F8-F10"),
    ("Financing", 1500000, 1800000, 2000000, 0, 0),
    ("Ending Cash", "=200000+B11+B12", "=B13+C11+C12", "=C13+D11+D12", "=D13+E11+E12", "=E13+F11+F12"),
]
for r in financial_rows:
    ws6.append(r)
style_header(ws6)
autofit(ws6, {"A": 32, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14})

# CAC payback
ws7 = wb.create_sheet("CAC_Payback")
ws7.append(["Year", "New Providers", "Sales+Marketing Spend", "CAC per Provider", "ARPU", "Monthly GP/Provider", "Payback Months"])
for i, row in enumerate(
    [
        # GTM spend includes performance marketing + attributable sales payroll/support
        ("Y1", 90, 180000, None, "=Growth_Pricing!K2", None, None),
        ("Y2", 230, 320000, None, "=Growth_Pricing!K3", None, None),
        ("Y3", 580, 700000, None, "=Growth_Pricing!K4", None, None),
        ("Y4", 1200, 1400000, None, "=Growth_Pricing!K5", None, None),
        ("Y5", 2100, 2700000, None, "=Growth_Pricing!K6", None, None),
    ],
    start=2,
):
    ws7.append(list(row))
    ws7[f"D{i}"] = f"=C{i}/B{i}"
    ws7[f"F{i}"] = f"=E{i}-COGS!$B$7"
    ws7[f"G{i}"] = f"=D{i}/F{i}"
style_header(ws7)
autofit(ws7, {"A": 8, "B": 14, "C": 22, "D": 17, "E": 10, "F": 18, "G": 15})

# SmartHelping mapping
ws8 = wb.create_sheet("SmartHelping_Mapping")
ws8.append(["Template Section", "Model Input / Formula Source"])
mapping_rows = [
    ("Monthly price / ARPU assumptions", "Growth_Pricing!K2:K6"),
    ("Subscriber growth assumptions", "Growth_Pricing!D2:D6"),
    ("New subscribers", "CAC_Payback!B2:B6"),
    ("Monthly churn", "Core_Assumptions!B4"),
    ("COGS per subscriber", "COGS!B7"),
    ("Marketing spend (non-payroll)", "Opex!B2:F2"),
    ("GTM spend for CAC payback", "CAC_Payback!C2:C6"),
    ("Payroll assumptions", "Hiring!A2:G9 and Hiring!B13:F13"),
    ("Other SG&A", "Opex!B8:F8"),
    ("Revenue", "Financials!B2:F2"),
    ("Gross margin", "COGS!F10:F14"),
    ("EBITDA", "Financials!B8:F8"),
    ("Cash flow and cash balance", "Financials!B11:F13"),
]
for r in mapping_rows:
    ws8.append(r)
style_header(ws8)
autofit(ws8, {"A": 34, "B": 48})

for sheet in wb.worksheets:
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=sheet.max_column):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"

out_path = "railguard_financial_model.xlsx"
wb.save(out_path)
print(f"Created {out_path}")
